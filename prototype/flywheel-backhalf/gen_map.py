#!/usr/bin/env python3
# 表现地图原型生成器：读平台 xlsx 导出 → 算「钩子(封面点击率/打开率) × 内容(收藏率)」漏斗 → 出可点 HTML
# 加公众号：拿到导出后在 load_wx() 填字段映射即可（打开率=阅读÷送达 当钩子力）
import openpyxl, statistics as st, json, math, os, glob
try: import xlrd
except Exception: xlrd=None

BASE = os.path.dirname(os.path.abspath(__file__))
DESK = os.path.expanduser("~/Desktop")
DOWN = os.path.expanduser("~/Downloads")

def num(x):
    try: return float(x)
    except: return 0.0

def is_ai(t):
    k=["AI","Agent","Claude","Codex","RAG","Skill","Token","Gemini","Harness","terminal","agent","协作",
       "自动","需求","文档","知识","收藏","提示词","脑子"]
    return any(x.lower() in (t or "").lower() for x in k)

def load_xhs(path):
    wb=openpyxl.load_workbook(path,data_only=True); ws=wb["Sheet1"]; out=[]
    for r in ws.iter_rows(min_row=3,values_only=True):
        title,pub,genre,exp,view,ctr,like,cmt,save,fans,share,dur,dm=r
        if exp is None: continue
        out.append(dict(plat="小红书",title=(title or "(无题)"),
            pub=str(pub)[:11].replace("年","-").replace("月","-").replace("日",""),genre=genre,
            imp=num(exp),view=num(view),ctr=num(ctr),like=num(like),cmt=num(cmt),
            save=num(save),fans=num(fans),share=num(share)))
    return out

def load_dy(path):
    wb=openpyxl.load_workbook(path,data_only=True); hdr=[c.value for c in next(wb["Sheet1"].iter_rows(min_row=1,max_row=1))]; out=[]
    for r in wb["Sheet1"].iter_rows(min_row=2,values_only=True):
        d=dict(zip(hdr,r))
        if d.get("播放量") is None: continue
        out.append(dict(plat="抖音",title=d["作品名称"],pub=str(d["发布时间"])[:10],genre=d["体裁"],
            imp=num(d["播放量"]),view=num(d["播放量"]),ctr=num(d.get("封面点击率")),like=num(d.get("点赞量")),
            cmt=num(d.get("评论量")),save=num(d.get("收藏量")),fans=num(d.get("粉丝增量")),share=num(d.get("分享量"))))
    return out

def load_wx(path):
    """公众号：推送模型。钩子力=打开率(阅读÷送达)填进 ctr 位；内容力=收藏率(收藏÷阅读)。
    待用户给真实导出后按实际列名映射。占位实现，找不到文件就跳过。"""
    if not path or not os.path.exists(path): return []
    wb=openpyxl.load_workbook(path,data_only=True); ws=wb.worksheets[0]
    hdr=[c.value for c in next(ws.iter_rows(min_row=1,max_row=1))]
    def col(*names):
        for n in names:
            if n in hdr: return hdr.index(n)
        return None
    i_t=col("标题","内容标题","图文标题"); i_pub=col("发表时间","群发时间","发布时间")
    i_send=col("送达人数","送达","推送人数"); i_read=col("阅读人数","阅读次数","阅读量","总阅读人数")
    i_save=col("收藏人数","收藏次数","收藏量","收藏"); i_share=col("分享人数","分享次数","分享")
    i_like=col("点赞","点赞人数","在看"); i_fans=col("涨粉","净增关注人数","新关注人数")
    out=[]
    for r in ws.iter_rows(min_row=2,values_only=True):
        if i_t is None or r[i_t] is None: continue
        send=num(r[i_send]) if i_send is not None else 0
        read=num(r[i_read]) if i_read is not None else 0
        save=num(r[i_save]) if i_save is not None else 0
        openrate=(read/send) if send else 0     # 打开率=钩子力
        out.append(dict(plat="公众号",title=r[i_t],pub=str(r[i_pub])[:10] if i_pub is not None else "",
            genre="图文",imp=send or read,view=read,ctr=openrate,
            like=num(r[i_like]) if i_like is not None else 0,cmt=0,
            save=save,fans=num(r[i_fans]) if i_fans is not None else 0,
            share=num(r[i_share]) if i_share is not None else 0))
    return out

def load_wx_detail(path):
    """公众号「单篇明细」.xls（纵向 指标|数值）。钩子力=订阅推送打开率=公众号消息阅读人数÷送达（老实的 [0,1]，
    区别于总阅读——总阅读含搜一搜/分享，会把打开率算爆）。一篇只出一个点。"""
    if not path or xlrd is None or not os.path.exists(path): return []
    sh=xlrd.open_workbook(path).sheets()[0]
    title=""; kv={}
    for i in range(sh.nrows):
        row=[sh.cell_value(i,c) for c in range(sh.ncols)]
        if i==0:
            title=next((str(c).strip() for c in row if str(c).strip()),"(公众号文章)")
        k=str(row[1]).strip() if len(row)>1 else ""
        if k and len(row)>2 and isinstance(row[2],(int,float)): kv[k]=float(row[2])
    read=kv.get("阅读(人)",0); send=kv.get("送达人数",0)
    push_read=kv.get("公众号消息阅读人数",0)
    hook = (push_read/send) if send else 0        # 订阅推送打开率
    return [dict(plat="公众号",title=title,pub="",genre="图文",
        imp=read, view=read, ctr=hook,
        like=kv.get("点赞(人)",0)+kv.get("在看(人)",0), cmt=kv.get("评论（条）",0),
        save=kv.get("收藏(人)",0), fans=kv.get("新增关注（人）",0), share=kv.get("总分享人数",kv.get("分享(人)",0)))]

# ---------- 装载 ----------
rows=[]
rows+=load_xhs(os.path.join(DESK,"小红书笔记列表明细表.xlsx"))
rows+=load_dy(os.path.join(DOWN,"抖音作品列表.xlsx"))
# 公众号：先找列表版 xlsx，再找单篇明细 .xls
wx=[]
for cand in [os.path.join(DOWN,"公众号数据.xlsx"),os.path.join(DESK,"公众号数据.xlsx"),
             os.path.join(DOWN,"微信公众号.xlsx"),os.path.join(DESK,"微信公众号.xlsx")]:
    wx=load_wx(cand)
    if wx: print(f"公众号列表已载入 {len(wx)} 条 ← {cand}"); break
if not wx:
    for g in glob.glob(os.path.join(DESK,"公众号数据明细*.xls"))+glob.glob(os.path.join(DOWN,"公众号数据明细*.xls")):
        d=load_wx_detail(g)
        if d: wx+=d; print(f"公众号单篇明细 +1 ← {os.path.basename(g)}")
if wx: rows+=wx
else: print("⚠ 未找到公众号导出")

rows=[x for x in rows if is_ai(x["title"])]
for x in rows:
    v=x["view"]; x["saveRate"]=x["save"]/v if v else 0
    x["engRate"]=(x["like"]+x["save"]+x["cmt"]+x["share"])/v if v else 0

# 基线：小红书样本足，用它定中位（公众号/抖音点落在同一坐标系里对照）
xhs=[x for x in rows if x["plat"]=="小红书"]
med_ctr=st.median([x["ctr"] for x in xhs if x["ctr"]>0])
med_sr =st.median([x["saveRate"] for x in xhs if x["view"]>=20])

def quad(x):
    hc=x["ctr"]>=med_ctr; hs=x["saveRate"]>=med_sr
    if hc and hs: return("双高·标杆","#3f7350")
    if hc and not hs: return("钩子强内容弱·修正文","#a9791f")
    if not hc and hs: return("内容强没被看见·换钩子重发","#3d5a80")
    return("双低·重做","#a24b3f")
for x in rows: x["quad"],x["qc"]=quad(x)

maxc=max([x["ctr"] for x in rows]+[0.01])*1.12
maxs=max([x["saveRate"] for x in rows]+[0.01])*1.12
maximp=max([x["imp"] for x in rows]+[1])
PL,PR,PT,PB=9,4,5,9
px=lambda v: PL+(v/maxc)*(100-PL-PR)
py=lambda v:(100-PB)-(v/maxs)*(100-PT-PB)
size=lambda imp:11+(math.log10(imp+1)/math.log10(maximp+1))*30
pts=[]
for x in rows:
    pts.append(dict(l=round(px(x["ctr"]),2),t=round(py(x["saveRate"]),2),s=round(size(x["imp"]),1),
        plat=x["plat"],qc=x["qc"],quad=x["quad"],title=x["title"],pub=x["pub"],
        ctr=round(x["ctr"],3),sr=round(x["saveRate"]*100,1),imp=int(x["imp"]),view=int(x["view"]),
        save=int(x["save"]),fans=int(x["fans"]),like=int(x["like"]),cmt=int(x["cmt"]),
        share=int(x["share"]),low=x["imp"]<500))
medx=round(px(med_ctr),2); medy=round(py(med_sr),2)
from collections import Counter
print(f"基线 钩子中位{med_ctr:.3f} 收藏中位{med_sr:.2%} | 点{len(pts)} | 平台",dict(Counter(p['plat'] for p in pts)))

DATA=json.dumps(pts,ensure_ascii=False)
TPL=r'''<!DOCTYPE html><html lang="zh"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>表现地图 · 真实数据（钩子 × 内容）</title>
<link href="https://fonts.googleapis.com/css2?family=Newsreader:opsz,wght@6..72,400;6..72,500;6..72,600&display=swap" rel="stylesheet">
<style>
:root{--bg:#efece4;--surface:#faf8f2;--brief:#f4f6f9;--text:#211f1a;--body:#332f27;--sub:#706b60;--sub2:#8a8478;--faint:#b0a894;--line10:rgba(33,31,26,.1);--line08:rgba(33,31,26,.08);--accent:#3d5a80;--amber:#a9791f;--green:#3f7350;--red:#a24b3f;--serif:'Newsreader',serif}
*{box-sizing:border-box}body{margin:0;background:var(--bg);color:var(--text);font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;font-size:14px;line-height:1.5}
.wrap{max-width:1140px;margin:0 auto;padding:24px 22px 90px}
h1{font-family:var(--serif);font-size:25px;font-weight:600;margin:0 0 4px}
.sub{font-size:13px;color:var(--sub2);margin-bottom:6px}
.note{font-size:12.5px;color:var(--sub);background:var(--surface);border:1px solid var(--line10);border-radius:10px;padding:12px 15px;line-height:1.7;margin:14px 0 20px}
.note b{color:var(--text)}
.main{display:flex;gap:22px;align-items:flex-start;flex-wrap:wrap}
.plot-card{flex:1;min-width:520px;background:var(--surface);border:1px solid var(--line10);border-radius:14px;padding:18px 20px 16px}
.plot-h{display:flex;justify-content:space-between;align-items:baseline;margin-bottom:12px}
.plot-h .t{font-family:var(--serif);font-size:17px;font-weight:600}
.filters{display:flex;gap:6px;flex-wrap:wrap}
.fbtn{border:1px solid var(--line10);background:var(--surface);color:var(--sub);font-size:11.5px;font-weight:600;padding:4px 10px;border-radius:7px;cursor:pointer}
.fbtn.on{background:var(--accent);color:#fff;border-color:var(--accent)}
.scatter{position:relative;width:100%;aspect-ratio:1/.82;background:linear-gradient(180deg,#f7f9fc,#f2f4f7);border:1px solid var(--line10);border-radius:10px}
.qbg{position:absolute;font-size:11px;font-weight:600;opacity:.9;pointer-events:none;line-height:1.3}
.crossv{position:absolute;top:2%;bottom:9%;border-left:1px dashed rgba(33,31,26,.22)}
.crossh{position:absolute;left:9%;right:4%;border-top:1px dashed rgba(33,31,26,.22)}
.medlab{position:absolute;font-size:9.5px;color:var(--sub2);background:rgba(255,255,255,.7);padding:0 3px;border-radius:3px}
.axx{position:absolute;bottom:2%;left:50%;transform:translateX(-50%);font-size:11px;color:var(--sub2);font-weight:600}
.axy{position:absolute;left:1%;top:50%;transform:translateY(-50%);writing-mode:vertical-rl;font-size:11px;color:var(--sub2);font-weight:600;letter-spacing:1px}
.dot{position:absolute;border-radius:50%;transform:translate(-50%,-50%);cursor:pointer;border:2px solid #fff;box-shadow:0 1px 5px rgba(0,0,0,.18);transition:transform .1s}
.dot:hover{transform:translate(-50%,-50%) scale(1.22);z-index:9}
.dot.low{opacity:.5;border-style:dashed}
.dot.dy{border-color:#211f1a}.dot.wx{border-color:#3f7350}
.dot.sel{outline:3px solid rgba(61,90,128,.35);outline-offset:1px}
#tip{position:fixed;pointer-events:none;z-index:99;background:rgba(28,26,22,.95);color:#fff;font-size:12px;padding:7px 10px;border-radius:8px;max-width:260px;line-height:1.45;box-shadow:0 4px 14px rgba(0,0,0,.25);display:none}
#tip .tt{font-weight:600;margin-bottom:3px}#tip .tm{font-size:10.5px;opacity:.8}
.leg{font-size:11px;color:var(--sub);margin-top:10px;line-height:1.7}
.leg .sw{display:inline-block;width:9px;height:9px;border-radius:50%;vertical-align:middle;margin:0 3px 0 10px;border:1.5px solid #fff}
.side{width:330px;flex:0 0 auto}
.detail{background:var(--surface);border:1px solid var(--line10);border-radius:14px;padding:16px 17px;position:sticky;top:16px}
.detail .badge{display:inline-block;font-size:11px;font-weight:600;padding:2px 10px;border-radius:20px;margin-bottom:9px}
.detail h3{font-family:var(--serif);font-size:16px;margin:0 0 3px;line-height:1.35}
.detail .meta{font-size:11.5px;color:var(--sub2);margin-bottom:12px}
.funnel{display:flex;gap:7px;margin:12px 0}
.fcell{flex:1;text-align:center;background:var(--brief);border:1px solid var(--line08);border-radius:8px;padding:8px 3px}
.fcell .v{font-size:15px;font-weight:700;font-variant-numeric:tabular-nums}.fcell .l{font-size:9.5px;color:var(--sub2);margin-top:2px}
.reco{font-size:12.5px;color:var(--body);background:var(--brief);border-left:3px solid var(--accent);border-radius:0 8px 8px 0;padding:10px 12px;margin-top:12px;line-height:1.6}.reco b{color:var(--accent)}
.raw{font-size:11px;color:var(--sub);margin-top:12px;line-height:1.8}.raw span{color:var(--body);font-weight:600}
.ranks{margin-top:20px}.ranks h4{font-family:var(--serif);font-size:14px;margin:0 0 8px}
.rline{display:flex;gap:8px;align-items:baseline;font-size:12px;padding:5px 0;border-bottom:1px solid var(--line08);cursor:pointer}
.rline:hover{background:rgba(61,90,128,.04)}
.rline .rt{flex:1;color:var(--body);overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.rline .rv{color:var(--sub2);font-variant-numeric:tabular-nums;font-size:11px}
</style></head><body><div class="wrap">
<h1>表现地图 · 你的真实内容</h1>
<div class="sub" id="subline">AI/科技内容 · 已滤掉美食生活 · 数据来自平台导出</div>
<div class="note">
<div style="font-weight:600;color:var(--text);margin-bottom:7px">看图说明</div>
<div style="display:grid;grid-template-columns:1fr 1fr;gap:6px 22px">
<div>■ <b>横轴 钩子力</b>：小红书/抖音＝封面点击率、公众号＝打开率——<b>能不能被点开</b>（标题+封面决定）</div>
<div>■ <b>纵轴 内容力（收藏率）</b>：收藏率＝收藏数/观看数，读者觉不觉得<b>有收获</b>（正文决定，你的第一尺子）</div>
<div>■ <b>点大小＝曝光/送达</b>：越大越可信</div>
<div>■ <b>虚线圈＝低样本</b>（曝光&lt;500）：别过度解读</div>
<div>■ <b>十字线＝你自己账号中位</b>：比的是"这篇比你平常强还是弱"</div>
<div>■ <b>鼠标悬停即显标题</b>，点击看完整诊断</div>
</div>
</div>
<div class="main">
 <div class="plot-card">
  <div class="plot-h"><span class="t">钩子 × 内容 四象限</span>
   <div class="filters"><button class="fbtn on" data-f="all">全部</button><button class="fbtn" data-f="小红书">小红书</button><button class="fbtn" data-f="抖音">抖音</button><button class="fbtn" data-f="公众号">公众号</button><button class="fbtn" data-f="hi">仅高样本</button></div>
  </div>
  <div class="scatter" id="sc">
   <div class="crossv" style="left:MEDXpx%"></div><div class="crossh" style="top:MEDYpx%"></div>
   <div class="medlab" style="left:MEDXpx%;top:3%;transform:translateX(4px)">钩子中位 MEDCTR</div>
   <div class="medlab" style="top:MEDYpx%;left:9.5%;transform:translateY(-14px)">收藏中位 MEDSR%</div>
   <div class="qbg" style="right:5%;top:6%;color:var(--green)">↗ 双高·标杆</div>
   <div class="qbg" style="left:11%;top:6%;color:var(--accent);text-align:left">内容强没被看见<br><span style="font-weight:400;color:var(--sub2)">换钩子·重发</span></div>
   <div class="qbg" style="right:5%;bottom:12%;color:var(--amber);text-align:right">钩子强内容弱<br><span style="font-weight:400;color:var(--sub2)">修正文</span></div>
   <div class="qbg" style="left:11%;bottom:12%;color:var(--red)">双低·重做</div>
   <div class="axx">钩子力（封面点击率/打开率）→</div><div class="axy">内容力（收藏率）→</div>
  </div>
  <div class="leg"><span class="sw" style="background:var(--accent);border-color:#fff"></span>小红书<span class="sw" style="background:#8a8478;border-color:#211f1a"></span>抖音<span class="sw" style="background:#8a8478;border-color:#3f7350"></span>公众号 · 点大=曝光高 · 填色=所在象限 · 虚线圈=低样本</div>
 </div>
 <div class="side"><div class="detail" id="dt">
   <div class="badge" style="background:rgba(33,31,26,.06);color:var(--sub2)">点一个点看诊断</div>
   <h3 id="dtT">你的内容落在这里</h3>
   <div class="meta" id="dtM">四象限＝四种处方。左上是"被低估的金子"，右下是"标题党"。</div>
   <div id="dtBody" style="display:none">
     <div class="funnel">
       <div class="fcell"><div class="v" id="fImp">–</div><div class="l" id="fImpL">曝光</div></div>
       <div class="fcell"><div class="v" id="fCtr">–</div><div class="l" id="fCtrL">钩子力</div></div>
       <div class="fcell"><div class="v" id="fSr">–</div><div class="l">收藏率</div></div>
       <div class="fcell"><div class="v" id="fFan">–</div><div class="l">涨粉</div></div>
     </div>
     <div class="reco" id="dtReco"></div><div class="raw" id="dtRaw"></div>
   </div>
   <div class="ranks"><h4>被低估的金子 · 内容强没被看见</h4><div id="goldList"></div></div>
 </div></div>
</div>
<div class="note" style="margin-top:22px"><b>这就是"极轻版"要验证的：看着能不能激发方向感？</b> 若行我落成产品——上传各平台 xlsx → 自动算漏斗入库 → 这张图 + 发布台账。归拢方式＝<b>每周下载平台导出丢进来</b>（不碰 cookie/API）。公众号=推送模型，用<b>打开率当钩子力</b>；深评·私信导出无、手动记。</div>
<div id="tip"><div class="tt"></div><div class="tm"></div></div>
<script>
const PTS=__DATA__;
const RECO={
 "双高·标杆":"<b>标杆。</b>钩子和内容都在你水平线之上。拆它的主题/角度/标题/开头，当模板复制到下一篇。",
 "钩子强内容弱·修正文":"<b>标题把人骗进来了，正文没接住。</b>钩子可复用，但要回炉正文——信息密度/收获感不够。",
 "内容强没被看见·换钩子重发":"<b>被低估的金子。</b>正文读者认可（收藏率高于中位），但钩子弱没拿到曝光。换更强标题/封面重发，很可能翻盘。",
 "双低·重做":"<b>钩子内容都低于平常。</b>选题或两头没做好——低样本先别急着下结论。"};
const PLABEL={"小红书":"封面点击率","抖音":"封面点击率","公众号":"打开率"};
const PIMP={"小红书":"曝光","抖音":"播放","公众号":"送达"};
const sc=document.getElementById('sc'),tip=document.getElementById('tip');
function render(f){
 sc.querySelectorAll('.dot').forEach(e=>e.remove());
 PTS.forEach((p,i)=>{
   if(f==='小红书'&&p.plat!=='小红书')return;if(f==='抖音'&&p.plat!=='抖音')return;
   if(f==='公众号'&&p.plat!=='公众号')return;if(f==='hi'&&p.low)return;
   const d=document.createElement('div');
   d.className='dot '+(p.plat==='抖音'?'dy ':p.plat==='公众号'?'wx ':'')+(p.low?'low':'');
   d.style.left=p.l+'%';d.style.top=p.t+'%';d.style.width=p.s+'px';d.style.height=p.s+'px';d.style.background=p.qc;
   d.onclick=()=>show(i,d);
   d.onmousemove=(e)=>{tip.style.display='block';tip.style.left=(e.clientX+14)+'px';tip.style.top=(e.clientY+14)+'px';
     tip.querySelector('.tt').textContent=p.title;
     tip.querySelector('.tm').textContent=p.plat+' · '+p.pub+' · '+PIMP[p.plat]+p.imp+(p.low?' · 低样本':'');};
   d.onmouseleave=()=>tip.style.display='none';
   sc.appendChild(d);
 });
}
function show(i,el){
 const p=PTS[i];document.querySelectorAll('.dot.sel').forEach(e=>e.classList.remove('sel'));if(el)el.classList.add('sel');
 document.getElementById('dtBody').style.display='block';
 const b=document.querySelector('#dt .badge');b.textContent=p.quad+(p.low?' · 低样本':'');
 b.style.background='color-mix(in srgb,'+p.qc+' 14%,transparent)';b.style.color=p.qc;
 document.getElementById('dtT').textContent=p.title;
 document.getElementById('dtM').textContent=p.plat+' · '+p.pub+' · '+p.imp+' '+PIMP[p.plat];
 document.getElementById('fImp').textContent=p.imp;document.getElementById('fImpL').textContent=PIMP[p.plat];
 document.getElementById('fCtr').textContent=p.ctr.toFixed(3);document.getElementById('fCtrL').textContent=PLABEL[p.plat];
 document.getElementById('fSr').textContent=p.sr+'%';document.getElementById('fFan').textContent='+'+p.fans;
 document.getElementById('dtReco').innerHTML=RECO[p.quad]+(p.low?' <span style="color:var(--amber)">⚠ 低样本，当趋势别当结论。</span>':'');
 document.getElementById('dtRaw').innerHTML='观看 <span>'+p.view+'</span> · 收藏 <span>'+p.save+'</span> · 赞 <span>'+p.like+'</span> · 评 <span>'+p.cmt+'</span> · 分享 <span>'+p.share+'</span>';
}
document.querySelectorAll('.fbtn').forEach(btn=>btn.onclick=()=>{document.querySelectorAll('.fbtn').forEach(b=>b.classList.remove('on'));btn.classList.add('on');render(btn.dataset.f);});
const gold=PTS.map((p,i)=>({p,i})).filter(o=>o.p.quad.includes('换钩子')).sort((a,b)=>b.p.sr-a.p.sr);
document.getElementById('goldList').innerHTML=gold.map(o=>`<div class="rline" onclick="show(${o.i})"><span class="rt">${o.p.title}</span><span class="rv">收藏${o.p.sr}% · 钩子${o.p.ctr.toFixed(3)}</span></div>`).join('')||'<div style="font-size:12px;color:var(--sub2)">暂无</div>';
document.getElementById('subline').textContent='共 '+PTS.length+' 篇 AI/科技内容（'+[...new Set(PTS.map(p=>p.plat))].join(' + ')+'）· 已滤掉美食生活 · 数据来自平台导出';
render('all');
</script></body></html>'''
out=(TPL.replace("__DATA__",DATA).replace("MEDXpx",str(medx)).replace("MEDYpx",str(medy))
   .replace("MEDCTR",f"{med_ctr:.3f}").replace("MEDSR",f"{med_sr*100:.1f}"))
open(os.path.join(BASE,"performance-map-real.html"),"w").write(out)
print("✅ performance-map-real.html 重生成")
